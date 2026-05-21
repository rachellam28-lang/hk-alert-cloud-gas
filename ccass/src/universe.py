"""Stock universe management.

CCASS 公開 stock list 喺 HKEX 網站。我哋每星期 refresh 一次。
"""
from __future__ import annotations

import time
from datetime import datetime
from typing import Iterable

import requests
from bs4 import BeautifulSoup

from src.db import get_conn
from src.logger import setup_logger

logger = setup_logger("universe")

# HKEX 有公開嘅 stock list CSV
HKEX_STOCK_LIST_URL = (
    "https://www.hkex.com.hk/eng/services/trading/securities/securitieslists/"
    "ListOfSecurities.xlsx"
)
HKEX_STOCK_LIST_CN_URL = (
    "https://www.hkex.com.hk/chi/services/trading/securities/securitieslists/"
    "ListOfSecurities_c.xlsx"
)


def fetch_all_hk_stocks_from_ccass() -> list[tuple[str, str]]:
    """
    Fallback: 從 CCASS query page 攞 dropdown stock list。
    回傳 [(stock_code_5digit, name), ...]

    我哋實際上唔需要逐隻 hardcode。每日 scrape 時，
    HKEX CCASS 系統會接受任何 5 位 stock code，invalid 嘅返回空。

    呢個 function 用 HKEX 嘅 stock list xlsx 作為 source of truth。
    會攞埋中文版 xlsx 嚟顯示中文公司名。
    """
    logger.info("Fetching HKEX stock list from %s", HKEX_STOCK_LIST_URL)
    last_err = None
    for attempt, backoff in enumerate((2, 4, 8), start=1):
        try:
            resp = requests.get(HKEX_STOCK_LIST_URL, timeout=60)
            resp.raise_for_status()
            break
        except requests.RequestException as e:
            last_err = e
            logger.warning("HKEX stock list attempt %d/3 failed: %s", attempt, e)
            if attempt < 3:
                time.sleep(backoff)
    else:
        logger.error("HKEX stock list failed after 3 attempts: %s", last_err)
        raise last_err

    # Parse xlsx — scan all sheets, pick the one with most stock codes
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content), data_only=True)
    logger.info("Sheets in xlsx: %s", wb.sheetnames)

    best_stocks: list[tuple[str, str]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        stocks: list[tuple[str, str]] = []
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            row_count += 1
            if row_count <= 5:
                logger.info("xlsx row %d: %s", row_count, [repr(c) for c in (row[:5] if row else [])])
            if not row or row[0] is None:
                continue
            raw = row[0]
            if isinstance(raw, float):
                if raw != int(raw):
                    continue
                raw = int(raw)
            if isinstance(raw, int):
                if raw <= 0 or raw > 99999:
                    continue
                code_str = str(raw)
            else:
                code_str = str(raw).strip().lstrip("'")
                if code_str.endswith('.0'):
                    code_str = code_str[:-2]
            if not code_str or not code_str[0].isdigit():
                continue
            if not (code_str.isdigit() and 1 <= len(code_str) <= 5):
                continue
            category = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            if category not in ("Equity", "Real Estate Investment Trusts", ""):
                continue
            code = code_str.zfill(5)
            name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            stocks.append((code, name))
        logger.info("Sheet '%s': %d rows total, %d stocks found", sheet_name, row_count, len(stocks))
        if len(stocks) > len(best_stocks):
            best_stocks = stocks

    # Fetch Chinese names from the Chinese xlsx
    try:
        cn_resp = requests.get(HKEX_STOCK_LIST_CN_URL, timeout=60)
        cn_resp.raise_for_status()
        cn_wb = load_workbook(io.BytesIO(cn_resp.content), data_only=True)
        cn_sheet = cn_wb.active
        cn_map: dict[str, str] = {}
        for row in cn_sheet.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            raw = row[0]
            if isinstance(raw, float):
                if raw != int(raw):
                    continue
                raw = int(raw)
            if isinstance(raw, int):
                if raw <= 0 or raw > 99999:
                    continue
                c_code = str(raw).zfill(5)
            else:
                c_code = str(raw).strip().lstrip("'")
                if c_code.endswith('.0'):
                    c_code = c_code[:-2]
                c_code = c_code.zfill(5)
            if not c_code.isdigit() or len(c_code) != 5:
                continue
            cn_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if cn_name and cn_name != "nan":
                cn_map[c_code] = cn_name
        # Overwrite English names with Chinese
        merged = 0
        for i, (code, en_name) in enumerate(best_stocks):
            if code in cn_map:
                best_stocks[i] = (code, cn_map[code])
                merged += 1
        logger.info("Merged %d/%d Chinese stock names", merged, len(best_stocks))
    except Exception as e:
        logger.warning("Chinese stock list fetch failed (English names kept): %s", e)

    logger.info("Found %d stocks total (best sheet)", len(best_stocks))
    return best_stocks


def refresh_universe() -> int:
    """Refresh stock_universe table. 回傳新增 stock 數量。"""
    stocks = fetch_all_hk_stocks_from_ccass()
    now_iso = datetime.utcnow().isoformat()
    added = 0

    with get_conn() as conn:
        for code, name in stocks:
            cur = conn.execute(
                "SELECT 1 FROM stock_universe WHERE stock_code = ?", (code,)
            )
            if cur.fetchone():
                conn.execute(
                    """UPDATE stock_universe
                       SET stock_name = ?, last_seen_at = ?, is_active = 1
                       WHERE stock_code = ?""",
                    (name, now_iso, code),
                )
            else:
                conn.execute(
                    """INSERT INTO stock_universe
                       (stock_code, stock_name, is_active, added_at, last_seen_at)
                       VALUES (?, ?, 1, ?, ?)""",
                    (code, name, now_iso, now_iso),
                )
                added += 1

    logger.info("Universe refreshed: %d added, %d total", added, len(stocks))
    return added


def get_active_stocks() -> list[str]:
    """攞所有 active stock codes。"""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT stock_code FROM stock_universe WHERE is_active = 1 ORDER BY stock_code"
        ).fetchall()
    return [r["stock_code"] for r in rows]


if __name__ == "__main__":
    refresh_universe()
